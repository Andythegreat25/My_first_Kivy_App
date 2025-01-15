from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from openpyxl import load_workbook, Workbook
from kivy.uix.label import Label
from datetime import datetime
from kivy.lang import Builder
from kivy.uix.popup import Popup

#caricare il file .kv
Builder.load_file("rifornimento.kv")

# Definizione del Popup di successo
Builder.load_string('''
<SuccessPopup@Popup>:
    title: 'Successo'
    size_hint: 0.6, 0.4
    auto_dismiss: True
    BoxLayout:
        orientation: 'vertical'
        Label:
            text: 'Rifornimento aggiunto con successo!'
        Button:
            text: 'OK'
            size_hint_y: None
            height: '40dp'
            on_press: root.dismiss()
''')

# aggiungere i campi per la registrazione del rifornimento
class RifornimentoApp(BoxLayout):
    def aggiungi_rifornimento(self, nome, cantiere, targa, fonte, destinazione, litri, km, note):
        file = "Carburante.xlsx"
        data = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        try:
            wb = load_workbook(file)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Data", "Nome Completo", "Cantiere", "Targa", "Fonte Gasolio", "Destinazione Gasolio", "Litri", "Km/Ore", "Note"])
        sheet.append([data, nome, cantiere, targa, fonte, destinazione, float(litri), float(km), note])
        wb.save(file)

         # Mostra il popup di successo
        success_popup = Popup(title='Congratulazioni!',
                              content=Label(text='Rifornimento aggiunto con successo!'),
                              size_hint=(0.6, 0.4),
                              auto_dismiss=True)
        success_popup.open()

    # aggiungere il metodo per visualizzare i dati
    def mostra_dati(self):
        file = "Carburante.xlsx"
        try:
            wb = load_workbook(file)
            sheet = wb.active
            dati = [row for row in sheet.iter_rows(values_only=True)]
            return dati
        except FileNotFoundError:
            return ["Il file non esiste! inserisci un rifornimento per crearlo"]

 # creare la classe principale
class RifornimentoAppMain(App):
    def build(self):
        self.icon = "LOGO VERTICALE.png"
        self.title = "APP CARBURANTE"
        return RifornimentoApp()

 # eseguire l'applicazione
if __name__ == "__main__":
    RifornimentoAppMain().run()
